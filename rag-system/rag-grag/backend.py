import os
import hashlib
import json
import re
import subprocess
import tempfile
import logging

import torch
import numpy as np
from pymilvus import Collection, connections, FieldSchema, CollectionSchema, DataType, Index, IndexType, utility
from neo4j import GraphDatabase
from halo import Halo
from PyPDF2 import PdfReader

# For chunking and sentence embeddings
from langchain.text_splitter import RecursiveCharacterTextSplitter
from sentence_transformers import SentenceTransformer

# Remove or comment out the old Huggingface imports if not used anymore
# from transformers import AutoTokenizer, AutoModel

# Optional file readers
try:
    import docx
except ImportError:
    docx = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

try:
    from PIL import Image
    import pytesseract
except ImportError:
    Image = None
    pytesseract = None

try:
    import whisper
except ImportError:
    whisper = None

# Optional progress bar
from tqdm import tqdm

###############################################################################
# Setup Logging
###############################################################################

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

###############################################################################
# Global variables
###############################################################################
collection = None
driver = None

# Old model references removed (tokenizer, model). We'll use a SentenceTransformer now.
embedding_model = None

###############################################################################
# Milvus, Neo4j, and Embedding Model Initialization
###############################################################################

def init_milvus_collection(
    host="localhost",
    port="19530",
    collection_name="document_embeddings"
):
    """
    Connects to Milvus and initializes the 'document_embeddings' collection
    with dimension 384 (for sentence-transformers/all-MiniLM-L6-v2).
    Adjust the dimension if you choose a different embedding model.
    """
    global collection
    connections.connect("default", host=host, port=port)
    logger.info(f"Connected to Milvus at {host}:{port}")

    # Check if collection exists
    if collection_name not in utility.list_collections():
        fields = [
            FieldSchema(
                name="document_id",
                dtype=DataType.INT64,
                is_primary=True,
                auto_id=False
            ),
            # Adjust this dim to match your chosen SentenceTransformer model
            FieldSchema(
                name="embedding",
                dtype=DataType.FLOAT_VECTOR,
                dim=384
            ),
        ]
        schema = CollectionSchema(fields, description="Document Embeddings")

        with Halo(
            text=f"Creating Milvus collection '{collection_name}' since it does not exist...",
            spinner="dots"
        ) as spinner:
            collection_local = Collection(
                name=collection_name,
                schema=schema
            )
            spinner.succeed(f"Collection '{collection_name}' created successfully.")

        index_params = {
            "index_type": IndexType.HNSW,
            "metric_type": "COSINE",
            "params": {"M": 16, "efConstruction": 200},
        }

        with Halo(text="Creating index on the embedding field...", spinner="dots") as spinner:
            Index(collection_local, "embedding", index_params)
            spinner.succeed("Index created successfully.")

        collection_local.load()
        collection = collection_local
    else:
        with Halo(text=f"Loading existing Milvus collection '{collection_name}'...", spinner="dots") as spinner:
            collection_local = Collection(collection_name)
            collection_local.load()
            spinner.succeed(f"Collection '{collection_name}' loaded successfully.")
        collection = collection_local

def init_neo4j(
    neo4j_uri="bolt://localhost:7687",
    neo4j_user="neo4j",
    neo4j_password="testtest"
):
    global driver
    driver = GraphDatabase.driver(neo4j_uri, auth=(neo4j_user, neo4j_password))
    logger.info(f"Connected to Neo4j at {neo4j_uri}")

def init_embedding_model(model_name="sentence-transformers/all-MiniLM-L6-v2"):
    """
    Initializes a Sentence-BERT or similar model from the sentence-transformers library.
    """
    global embedding_model
    logger.info(f"Loading embedding model '{model_name}'...")
    with Halo(text="Loading embedding model...", spinner="dots") as spinner:
        embedding_model = SentenceTransformer(model_name)
        spinner.succeed(f"Embedding model '{model_name}' loaded successfully.")

def initialize_all(
    neo4j_uri="bolt://localhost:7687",
    neo4j_user="neo4j",
    neo4j_password="testtest",
    milvus_host="localhost",
    milvus_port="19530"
):
    """
    Call this once to set up Milvus, Neo4j, and the embedding model.
    """
    logger.info("Initializing all services (Milvus, Neo4j, Embedding Model)...")
    init_milvus_collection(host=milvus_host, port=milvus_port, collection_name="document_embeddings")
    init_neo4j(neo4j_uri=neo4j_uri, neo4j_user=neo4j_user, neo4j_password=neo4j_password)
    init_embedding_model()
    logger.info("All services initialized successfully.")

###############################################################################
# Neo4j / Graph Helpers
###############################################################################

def clear_neo4j_graph(batch_size=10):
    with driver.session() as session:
        with Halo(text="Clearing Neo4j graph in batches...", spinner="dots") as spinner:
            deleted = batch_size
            total_deleted = 0
            while deleted == batch_size:
                result = session.run(f"""
                MATCH (n)
                WITH n LIMIT {batch_size}
                DETACH DELETE n
                RETURN count(*) AS count
                """)
                record = result.single()
                deleted = record["count"] if record else 0
                total_deleted += deleted
                if deleted == 0:
                    break
            spinner.succeed(f"Neo4j graph cleared successfully. Total nodes deleted: {total_deleted}")

def _create_document_node_tx(tx, doc_id, content, metadata_json):
    tx.run(
        """
        CREATE (d:Document {
            doc_id: $doc_id,
            content: $content,
            metadata: $metadata_json,
            is_global: true,
            workspace_id: null
        })
        """,
        doc_id=doc_id,
        content=content,
        metadata_json=metadata_json
    )

def create_document_node(doc_id, content, metadata):
    metadata_json = json.dumps(metadata)
    with driver.session() as session:
        session.write_transaction(_create_document_node_tx, doc_id, content, metadata_json)
    logger.debug(f"Created Neo4j node for doc_id: {doc_id}")

def _create_relationship_tx(tx, doc_id_1, doc_id_2, relationship_type, extra_data):
    tx.run(
        """
        MATCH (d1:Document {doc_id: $doc_id_1}), (d2:Document {doc_id: $doc_id_2})
        CREATE (d1)-[:RELATED {type: $relationship_type, extra: $extra_data}]->(d2)
        """,
        doc_id_1=doc_id_1,
        doc_id_2=doc_id_2,
        relationship_type=relationship_type,
        extra_data=json.dumps(extra_data or {})
    )

def create_relationship(doc_id_1, doc_id_2, relationship_type, extra_data=None):
    if extra_data and "score" in extra_data:
        extra_data["score"] = float(extra_data["score"])
    with driver.session() as session:
        session.write_transaction(_create_relationship_tx, doc_id_1, doc_id_2, relationship_type, extra_data or {})
    logger.debug(f"Created relationship '{relationship_type}' between {doc_id_1} and {doc_id_2}")

###############################################################################
# Embedding & Similarity
###############################################################################

def generate_doc_id(content):
    sha256_hash = hashlib.sha256(content.encode('utf-8')).hexdigest()
    doc_id = int(sha256_hash, 16) % (2 ** 63 - 1)
    return doc_id

def store_embeddings(doc_ids, embeddings):
    """
    Stores the doc_ids and embeddings in Milvus. Ensures embeddings is float array.
    """
    embeddings = np.array(embeddings, dtype=np.float32)
    collection.insert([doc_ids, embeddings.tolist()])
    collection.flush()
    logger.debug(f"Stored {len(doc_ids)} embeddings in Milvus.")

def cosine_similarity(embedding1, embedding2):
    return np.dot(embedding1, embedding2) / (np.linalg.norm(embedding1) * np.linalg.norm(embedding2))

def compute_similarity_and_store(doc_id, embedding, doc_ids, embeddings, threshold=0.7):
    """
    For each embedding in embeddings, if the sim score > threshold, create a SIMILAR_TO edge.
    """
    for i, existing_embedding in enumerate(embeddings):
        existing_doc_id = doc_ids[i]
        similarity_score = cosine_similarity(embedding, existing_embedding)
        if similarity_score > threshold:
            create_relationship(doc_id, existing_doc_id, "SIMILAR_TO", extra_data={"score": similarity_score})
            logger.debug(f"Similarity {similarity_score:.4f} > {threshold} between {doc_id} and {existing_doc_id}")

###############################################################################
# File Reading & Utilities
###############################################################################

def transcribe_video(file_path, whisper_model_size="base"):
    """
    Transcribe video to text using ffmpeg + Whisper (if installed).
    """
    if not whisper:
        logger.warning("Whisper is not installed. Please install with `pip install openai-whisper`.")
        return None

    with tempfile.TemporaryDirectory() as tmpdir:
        audio_path = os.path.join(tmpdir, "temp_audio.wav")

        ffmpeg_cmd = [
            "ffmpeg",
            "-i", file_path,
            "-ar", "16000",  # 16kHz
            "-ac", "1",      # mono
            "-f", "wav",
            "-y",            # overwrite
            audio_path
        ]

        try:
            subprocess.run(ffmpeg_cmd, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            logger.debug(f"Extracted audio from {file_path} to {audio_path}")
        except subprocess.CalledProcessError as e:
            logger.error(f"Error extracting audio from video {file_path}: {e}")
            return None

        model = whisper.load_model(whisper_model_size)
        logger.info(f"Transcribing video {file_path} with Whisper model '{whisper_model_size}'...")
        result = model.transcribe(audio_path)
        transcription = result["text"].strip() if "text" in result else None
        logger.debug(f"Transcription result: {transcription[:60]}...")  # Show first 60 chars
        return transcription

def read_file_content(file_path):
    """
    Reads the file content into a single string, for further chunking or processing.
    """
    ext = os.path.splitext(file_path)[1].lower()

    try:
        if ext == ".md":
            with open(file_path, 'r', encoding='utf-8') as f:
                logger.debug(f"Reading Markdown file: {file_path}")
                return f.read()

        elif ext == ".pdf":
            try:
                reader = PdfReader(file_path)
                text = "\n".join([page.extract_text() for page in reader.pages if page.extract_text()])
                logger.debug(f"Extracted text from PDF file: {file_path}")
                return text
            except Exception as e:
                logger.error(f"Error reading PDF file {file_path}: {e}")
                return None

        elif ext == ".docx" and docx:
            try:
                doc = docx.Document(file_path)
                text = "\n".join([p.text for p in doc.paragraphs])
                logger.debug(f"Extracted text from DOCX file: {file_path}")
                return text
            except Exception as e:
                logger.error(f"Error reading DOCX file {file_path}: {e}")
                return None

        elif ext in [".xlsx", ".xls"] and load_workbook:
            try:
                wb = load_workbook(filename=file_path)
                text_content = []
                for sheet in wb:
                    for row in sheet.iter_rows(values_only=True):
                        row_text = " ".join([str(cell) for cell in row if cell is not None])
                        text_content.append(row_text)
                logger.debug(f"Extracted text from Spreadsheet file: {file_path}")
                return "\n".join(text_content)
            except Exception as e:
                logger.error(f"Error reading XLSX file {file_path}: {e}")
                return None

        elif ext in [".png", ".jpg", ".jpeg", ".tif", ".tiff"] and Image and pytesseract:
            try:
                image = Image.open(file_path)
                text = pytesseract.image_to_string(image)
                logger.debug(f"Extracted text from Image file: {file_path}")
                return text.strip() if text else None
            except Exception as e:
                logger.error(f"Error reading Image file {file_path}: {e}")
                return None

        elif ext in [".mp4", ".mov", ".avi", ".mkv"]:
            transcription = transcribe_video(file_path, whisper_model_size="base")
            return transcription

        else:
            # Try to read as text
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    logger.debug(f"Reading text file: {file_path}")
                    return f.read()
            except Exception as e:
                logger.error(f"Error reading text file {file_path}: {e}")
                return None

    except Exception as e:
        logger.error(f"An error occurred while reading {file_path}: {e}")
        return None

def extract_metadata(file_path, content):
    """
    Basic metadata for each file (or chunk).
    """
    metadata = {
        "filename": os.path.basename(file_path),
        "size": os.path.getsize(file_path),
        "word_count": len(content.split()) if content else 0,
        "links": []
    }

    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".md":
        metadata["type"] = "Markdown"
    elif ext == ".pdf":
        metadata["type"] = "PDF"
    elif ext == ".docx":
        metadata["type"] = "DOCX"
    elif ext in [".xlsx", ".xls"]:
        metadata["type"] = "Spreadsheet"
    elif ext in [".png", ".jpg", ".jpeg", ".tif", ".tiff"]:
        metadata["type"] = "Image"
    elif ext in [".mp4", ".mov", ".avi", ".mkv"]:
        metadata["type"] = "Video"
    else:
        metadata["type"] = "Text"
    return metadata

def extract_links_and_references(content, doc_map):
    """
    Example: we look for markdown links [text](url).
    doc_map can be used to find references to existing docs.
    """
    links = re.findall(r"\[.*?\]\((.*?)\)", content)
    relationships = []
    for link in links:
        for ref_doc, ref_doc_id in doc_map.items():
            if link in ref_doc:
                relationships.append((ref_doc_id, "LINK"))
    return links, relationships

###############################################################################
# CHUNKING & PROCESS
###############################################################################

def chunk_text(text, chunk_size=1000, chunk_overlap=200):
    """
    Uses LangChain's RecursiveCharacterTextSplitter to chunk large docs.
    """
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=chunk_size,
        chunk_overlap=chunk_overlap
    )
    chunks = splitter.split_text(text)
    logger.debug(f"Chunked text into {len(chunks)} chunks.")
    return chunks

def process_batch(docs_batch):
    """
    docs_batch: list of tuples (chunk_content, doc_id, metadata)
    We'll embed each chunk, create Neo4j nodes, store embeddings, etc.
    """
    embeddings = []
    doc_ids = []
    doc_map = {}

    chunk_texts = [item[0] for item in docs_batch]

    with Halo(text="Embedding batch of documents...", spinner="dots") as spinner:
        try:
            batch_embeddings = embedding_model.encode(chunk_texts, show_progress_bar=False)
            spinner.succeed("Batch embedding completed.")
        except Exception as e:
            spinner.fail(f"Failed to embed batch: {e}")
            return

    for i, (chunk_content, doc_id, metadata) in enumerate(docs_batch):
        # Create node in Neo4j
        create_document_node(doc_id, chunk_content, metadata)

        # Save for similarity linking
        embeddings.append(batch_embeddings[i])
        doc_ids.append(doc_id)

        # Optionally handle references in chunk
        # links, _ = extract_links_and_references(chunk_content, doc_map)
        # metadata["links"] = [link.split("/")[-1] for link in links]
        # doc_map[...] = doc_id

    # Compute similarities and create relationships
    logger.info(f"Computing similarities for batch of {len(doc_ids)} documents...")
    with Halo(text="Computing similarities and creating relationships...", spinner="dots") as spinner:
        try:
            for i, embedding in enumerate(embeddings):
                compute_similarity_and_store(doc_ids[i], embedding, doc_ids[:i], embeddings[:i])
            spinner.succeed("Similarity computation and relationship creation completed.")
        except Exception as e:
            spinner.fail(f"Failed during similarity computation: {e}")

    # Store embeddings in Milvus
    logger.info(f"Storing embeddings for batch of {len(doc_ids)} documents in Milvus...")
    try:
        store_embeddings(doc_ids, embeddings)
        logger.info("Embeddings stored in Milvus successfully.")
    except Exception as e:
        logger.error(f"Failed to store embeddings in Milvus: {e}")

def process_documents(directory, batch_size=50):
    """
    1. Clears the Neo4j graph (fresh start).
    2. Reads each file, chunks it.
    3. Batches the chunks into Milvus + Neo4j.
    """
    clear_neo4j_graph()

    all_docs = []
    logger.info(f"Reading and chunking documents from directory: {directory}")
    for filename in tqdm(os.listdir(directory), desc="Scanning files"):
        file_path = os.path.join(directory, filename)
        if not os.path.isfile(file_path):
            continue  # Skip subdirectories or non-files

        content = read_file_content(file_path)
        if not content:
            logger.warning(f"No content extracted from {file_path}. Skipping.")
            continue

        # Split into smaller chunks
        chunks = chunk_text(content, chunk_size=1000, chunk_overlap=200)
        metadata = extract_metadata(file_path, content)

        # For each chunk, generate a unique doc_id
        for chunk in chunks:
            doc_id = generate_doc_id(chunk)
            # We store chunk-level metadata (optionally you can add "chunk_index" or "filename" again).
            all_docs.append((chunk, doc_id, dict(metadata)))  # copy or dict() to avoid mutation

    logger.info(f"Total chunks to process: {len(all_docs)}")

    # Now process in batches with progress bar
    for i in tqdm(range(0, len(all_docs), batch_size), desc="Processing batches"):
        batch = all_docs[i:i+batch_size]
        process_batch(batch)

###############################################################################
# MAIN (optional usage)
###############################################################################

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Process and embed documents.")
    parser.add_argument(
        "--directory",
        type=str,
        default="/home/marek/rag-documents/global/",
        help="Directory containing documents to process."
    )
    parser.add_argument(
        "--batch_size",
        type=int,
        default=50,
        help="Number of documents to process in each batch."
    )
    parser.add_argument(
        "--log_level",
        type=str,
        default="INFO",
        help="Logging level (DEBUG, INFO, WARNING, ERROR, CRITICAL)."
    )
    args = parser.parse_args()

    # Set logging level based on user input
    numeric_level = getattr(logging, args.log_level.upper(), None)
    if not isinstance(numeric_level, int):
        logger.error(f'Invalid log level: {args.log_level}')
        exit(1)
    logger.setLevel(numeric_level)

    directory = args.directory
    if os.path.isdir(directory):
        initialize_all()
        with Halo(text="Starting document processing...", spinner="dots") as spinner:
            try:
                process_documents(directory, batch_size=args.batch_size)
                spinner.succeed("Document processing completed successfully.")
            except Exception as e:
                spinner.fail(f"An error occurred during processing: {e}")
    else:
        logger.error(f"Invalid directory: {directory}")
